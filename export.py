"""
Excel export module for generating expense reports.
"""

from io import BytesIO
from datetime import date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_report(expenses: list) -> BytesIO:
    """
    Generate an Excel report with all expenses and summary.

    Args:
        expenses: List of Expense objects

    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()

    # Create sheets
    ws_expenses = wb.active
    ws_expenses.title = "All Expenses"
    ws_summary = wb.create_sheet("Summary")

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    currency_alignment = Alignment(horizontal="right")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === Sheet 1: All Expenses ===
    headers = [
        "Date", "Type", "Category", "Vendor", "Explanation",
        "Amount", "Currency", "Amount (EUR)", "Exchange Rate",
        "Invoice #", "Tags", "Source"
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_expenses.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write expense data
    for row, expense in enumerate(expenses, 2):
        data = [
            expense.expense_date.isoformat() if expense.expense_date else "",
            expense.type.capitalize() if expense.type else "",
            expense.cost_category.capitalize() if expense.cost_category else "",
            expense.vendor_name or "",
            expense.explanation or "",
            float(expense.amount) if expense.amount else 0,
            expense.currency or "",
            float(expense.amount_eur) if expense.amount_eur else 0,
            float(expense.exchange_rate) if expense.exchange_rate else "",
            expense.invoice_number or "",
            ", ".join(expense.tags) if expense.tags else "",
            expense.source_type or ""
        ]

        for col, value in enumerate(data, 1):
            cell = ws_expenses.cell(row=row, column=col, value=value)
            cell.border = thin_border
            # Right-align currency columns
            if col in [6, 8, 9]:
                cell.alignment = currency_alignment
                if col in [6, 8] and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif col == 9 and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.000000'

    # Auto-adjust column widths
    column_widths = [12, 10, 12, 20, 30, 12, 10, 14, 14, 15, 20, 10]
    for col, width in enumerate(column_widths, 1):
        ws_expenses.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws_expenses.freeze_panes = "A2"

    # === Sheet 2: Summary ===
    # Calculate totals
    total_income = sum(
        float(e.amount_eur or 0) for e in expenses if e.type == 'income'
    )
    total_costs = sum(
        float(e.amount_eur or 0) for e in expenses if e.type == 'cost'
    )
    net = total_income - total_costs

    # Category breakdown
    category_totals = {
        'operations': Decimal('0'),
        'freelancers': Decimal('0'),
        'equipment': Decimal('0'),
        'other': Decimal('0'),
        'uncategorized': Decimal('0')
    }

    for e in expenses:
        if e.type == 'cost' and e.amount_eur:
            category = e.cost_category or 'uncategorized'
            if category in category_totals:
                category_totals[category] += Decimal(str(e.amount_eur))
            else:
                category_totals['uncategorized'] += Decimal(str(e.amount_eur))

    # Write summary
    summary_data = [
        ("Expense Summary", ""),
        ("", ""),
        ("Total Income (EUR)", total_income),
        ("Total Costs (EUR)", total_costs),
        ("Net (EUR)", net),
        ("", ""),
        ("Costs by Category", ""),
        ("Operations", float(category_totals['operations'])),
        ("Freelancers", float(category_totals['freelancers'])),
        ("Equipment", float(category_totals['equipment'])),
        ("Other", float(category_totals['other'])),
        ("Uncategorized", float(category_totals['uncategorized'])),
        ("", ""),
        ("Report Generated", date.today().isoformat()),
        ("Total Records", len(expenses))
    ]

    for row, (label, value) in enumerate(summary_data, 1):
        label_cell = ws_summary.cell(row=row, column=1, value=label)
        value_cell = ws_summary.cell(row=row, column=2, value=value)

        # Style headers
        if row == 1 or label == "Costs by Category":
            label_cell.font = Font(bold=True, size=14)
        elif label in ["Total Income (EUR)", "Total Costs (EUR)", "Net (EUR)"]:
            label_cell.font = Font(bold=True)
            value_cell.font = Font(bold=True)
            value_cell.number_format = '#,##0.00'
        elif label in ["Operations", "Freelancers", "Equipment", "Other", "Uncategorized"]:
            value_cell.number_format = '#,##0.00'

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def get_export_filename() -> str:
    """
    Generate filename for the export with current date.

    Returns:
        Filename string like "expenses_2024-01-15.xlsx"
    """
    return f"expenses_{date.today().isoformat()}.xlsx"
